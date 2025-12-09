# =======================================================================
# i3T4AN (Ethan Blair)
# Project:      Vector Knowledge Base
# File:         Excel spreadsheet text extractor
# =======================================================================

import io
from typing import Dict, Any, Tuple, Optional
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from .base import BaseExtractor
from exceptions import ExtractionError

class XlsxExtractor(BaseExtractor):
    """Extractor for Excel (.xlsx) files."""

    def extract(self, file_path: str, file_content: Optional[bytes] = None) -> Tuple[str, Dict[str, Any]]:
        """
        Extract text content and metadata from an Excel file.
        
        Args:
            file_path: Path to the file
            file_content: Optional binary content of the file
            
        Returns:
            Tuple containing (extracted_text, metadata)
        """
        try:
            if file_content:
                wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True, read_only=True)
            else:
                wb = openpyxl.load_workbook(filename=file_path, data_only=True, read_only=True)
            
            text_parts = []
            total_cells = 0
            
            # Extract metadata
            metadata = {
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "author": wb.properties.creator,
                "title": wb.properties.title,
                "created": str(wb.properties.created) if wb.properties.created else None,
                "modified": str(wb.properties.modified) if wb.properties.modified else None,
            }
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = []
                
                # Iterate through rows
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to string
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        sheet_text.append(" | ".join(row_values))
                        total_cells += len(row_values)
                
                if sheet_text:
                    text_parts.append(f"Sheet: {sheet_name}")
                    text_parts.append("\n".join(sheet_text))
                    text_parts.append("-" * 20)  # Separator between sheets
            
            metadata["total_cells_processed"] = total_cells
            
            # Close workbook
            wb.close()
            
            return "\n".join(text_parts), metadata
            
        except InvalidFileException as e:
            raise ExtractionError(f"Invalid Excel file: {str(e)}")
        except Exception as e:
            raise ExtractionError(f"Error extracting text from Excel file: {str(e)}")
